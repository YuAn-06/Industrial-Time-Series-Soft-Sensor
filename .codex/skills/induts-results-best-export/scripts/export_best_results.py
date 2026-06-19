#!/usr/bin/env python
"""Export best InduTS-SS result runs per model and collect artifacts."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np


METRIC_NAMES = ["mae", "mse", "rmse", "mape", "mspe", "wape", "corr_or_r2"]
HIGHER_IS_BETTER = {"corr", "r2", "corr_or_r2"}
TASK_ALIASES = {
    "SS_task": "soft_sensor",
    "ss_task": "soft_sensor",
    "ss": "soft_sensor",
    "soft_sensor": "soft_sensor",
    "LSF_task": "short_term_forecasting",
    "lsf_task": "short_term_forecasting",
    "lsf": "short_term_forecasting",
    "forecasting": "short_term_forecasting",
    "short_term_forecasting": "short_term_forecasting",
}
TASK_TO_SCRIPT_DIR = {
    "soft_sensor": "SS_task",
    "short_term_forecasting": "LSF_task",
}
PARAM_PATTERNS = {
    "seq_len": r"(?:^|_)sl(?P<value>-?\d+)",
    "label_len": r"(?:^|_)ll(?P<value>-?\d+)",
    "pred_len": r"(?:^|_)pl(?P<value>-?\d+)",
    "batch_size": r"(?:^|_)bt(?P<value>-?\d+)",
    "learning_rate": r"(?:^|_)lr(?P<value>-?\d+(?:\.\d+)?)",
    "epoch": r"(?:^|_)ep(?P<value>-?\d+)",
    "patience": r"(?:^|_)pat(?P<value>-?\d+)",
    "d_model": r"(?:^|_)dm(?P<value>-?\d+)",
    "d_ff": r"(?:^|_)dff(?P<value>-?\d+)",
    "n_heads": r"(?:^|_)nh(?P<value>-?\d+)",
    "e_layers": r"(?:^|_)el(?P<value>-?\d+)",
    "d_layers": r"(?:^|_)dl(?P<value>-?\d+)",
    "moving_avg": r"(?:^|_)ma(?P<value>-?\d+)",
    "individual": r"(?:^|_)ind(?P<value>True|False)",
    "kernel_size": r"(?:^|_)ks(?P<value>-?\d+)",
    "graph_threshold": r"(?:^|_)gt(?P<value>-?\d+(?:\.\d+)?)",
}
COPY_PATTERNS = [
    "metrics.*",
    "pred.*",
    "true.*",
    "test.*",
    "*.log",
    "*.yaml",
    "*.yml",
    "args.*",
    "config.*",
    "*.json",
]
CHECKPOINT_PATTERNS = ["checkpoint*.pth", "*.ckpt", "*.pt"]


def normalize_task(value: str | None) -> str | None:
    if not value:
        return None
    return TASK_ALIASES.get(value, TASK_ALIASES.get(value.lower(), value))


def parse_scalar(value: str) -> Any:
    if value in {"True", "False"}:
        return value == "True"
    try:
        if "." in value:
            return float(value)
        return int(value)
    except ValueError:
        return value


def parse_setting(model: str, setting: str) -> dict[str, Any]:
    task = None
    dataset = None
    marker = None
    for candidate in ("short_term_forecasting", "soft_sensor"):
        marker_candidate = f"_{model}_{candidate}"
        if marker_candidate in setting:
            task = candidate
            marker = marker_candidate
            dataset = setting.split(marker_candidate, 1)[0]
            break
    if task is None:
        for candidate in ("short_term_forecasting", "soft_sensor"):
            if candidate in setting:
                task = candidate
                dataset = setting.split("_", 1)[0]
                break
    row: dict[str, Any] = {
        "dataset": dataset or "",
        "task": task or "",
        "setting": setting,
    }
    suffix = setting
    if marker:
        suffix = setting.split(marker, 1)[1]
    for key, pattern in PARAM_PATTERNS.items():
        match = re.search(pattern, suffix)
        if match:
            row[key] = parse_scalar(match.group("value"))
    return row


def load_metrics(run_dir: Path) -> tuple[dict[str, float], str]:
    metrics_path = run_dir / "metrics.npy"
    if not metrics_path.exists():
        return {}, "missing_metrics"
    try:
        values = np.load(metrics_path, allow_pickle=True)
        flat = np.asarray(values, dtype=float).reshape(-1)
    except Exception as exc:  # noqa: BLE001
        return {}, f"metrics_error:{exc}"
    metrics = {}
    for name, value in zip(METRIC_NAMES, flat):
        if math.isfinite(float(value)):
            metrics[name] = float(value)
    return metrics, "ok" if metrics else "empty_metrics"


def discover_runs(results_dir: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not results_dir.exists():
        return rows
    for model_dir in sorted(path for path in results_dir.iterdir() if path.is_dir()):
        for run_dir in sorted(path for path in model_dir.iterdir() if path.is_dir()):
            setting_info = parse_setting(model_dir.name, run_dir.name)
            metrics, status = load_metrics(run_dir)
            rows.append(
                {
                    "model": model_dir.name,
                    **setting_info,
                    **metrics,
                    "status": status,
                    "run_dir": str(run_dir),
                }
            )
    return rows


def matches_filter(row: dict[str, Any], args: argparse.Namespace) -> bool:
    task = normalize_task(args.task)
    if task and row.get("task") != task:
        return False
    if args.dataset and str(row.get("dataset", "")).lower() != args.dataset.lower():
        return False
    if args.model and row.get("model") not in set(args.model):
        return False
    for arg_name, row_name in (
        ("seq_len", "seq_len"),
        ("label_len", "label_len"),
        ("pred_len", "pred_len"),
    ):
        expected = getattr(args, arg_name)
        if expected is not None and row.get(row_name) != expected:
            return False
    if args.setting_contains and args.setting_contains not in row.get("setting", ""):
        return False
    return True


def metric_direction(metric: str, requested: str) -> str:
    if requested != "auto":
        return requested
    if metric.lower() in HIGHER_IS_BETTER:
        return "max"
    return "min"


def select_best(rows: list[dict[str, Any]], metric: str, direction: str) -> list[dict[str, Any]]:
    best: list[dict[str, Any]] = []
    for model in sorted({row["model"] for row in rows}):
        candidates = [
            row for row in rows
            if row["model"] == model and isinstance(row.get(metric), (int, float))
        ]
        if not candidates:
            continue
        candidates.sort(key=lambda row: row[metric], reverse=(direction == "max"))
        selected = dict(candidates[0])
        selected["rank_metric"] = metric
        selected["rank_direction"] = direction
        selected["metric_value"] = selected[metric]
        best.append(selected)
    return best


def safe_name(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", value).strip("_") or "run"


def unique_destination(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    for index in range(1, 1000):
        candidate = parent / f"{stem}_{index}{suffix}"
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"Could not create unique file name for {path}")


def copy_matching(run_dir: Path, artifact_dir: Path, include_checkpoints: bool) -> list[str]:
    artifact_dir.mkdir(parents=True, exist_ok=True)
    copied: list[str] = []
    patterns = list(COPY_PATTERNS)
    if include_checkpoints:
        patterns.extend(CHECKPOINT_PATTERNS)
    seen: set[Path] = set()
    for pattern in patterns:
        for source in sorted(run_dir.glob(pattern)):
            if source in seen or not source.is_file():
                continue
            seen.add(source)
            destination = unique_destination(artifact_dir / source.name)
            shutil.copy2(source, destination)
            copied.append(str(destination))
    return copied


def copy_source_yaml(repo_root: Path, row: dict[str, Any], artifact_dir: Path) -> str:
    task_dir = TASK_TO_SCRIPT_DIR.get(str(row.get("task", "")))
    dataset = row.get("dataset")
    model = row.get("model")
    if not task_dir or not dataset or not model:
        return ""
    source = repo_root / "scripts" / task_dir / f"{dataset}_scripts" / "yaml" / f"{model}.yaml"
    if not source.exists():
        return ""
    destination = artifact_dir / f"source_{source.name}"
    shutil.copy2(source, destination)
    return str(destination)


def collect_artifacts(
    repo_root: Path,
    best_rows: list[dict[str, Any]],
    output_dir: Path,
    include_checkpoints: bool,
) -> list[dict[str, Any]]:
    manifest_rows = []
    artifact_root = output_dir / "best_artifacts"
    for row in best_rows:
        run_dir = Path(row["run_dir"])
        artifact_dir = artifact_root / safe_name(str(row["model"]))
        copied = copy_matching(run_dir, artifact_dir, include_checkpoints)
        source_yaml = copy_source_yaml(repo_root, row, artifact_dir)
        if source_yaml:
            copied.append(source_yaml)
        row["artifact_dir"] = str(artifact_dir)
        row["copied_file_count"] = len(copied)
        row["copied_files"] = "; ".join(copied)
        manifest_rows.append(
            {
                "model": row["model"],
                "run_dir": row["run_dir"],
                "artifact_dir": str(artifact_dir),
                "copied_files": copied,
                "missing_checkpoint": include_checkpoints
                and not any(Path(path).suffix in {".pth", ".ckpt", ".pt"} for path in copied),
                "missing_prediction": not any(Path(path).stem.startswith("pred") for path in copied),
                "missing_truth": not any(Path(path).stem.startswith("true") for path in copied),
            }
        )
    return manifest_rows


def fieldnames_for(rows: list[dict[str, Any]]) -> list[str]:
    preferred = [
        "model",
        "dataset",
        "task",
        "rank_metric",
        "rank_direction",
        "metric_value",
        *METRIC_NAMES,
        "status",
        "setting",
        "run_dir",
        "artifact_dir",
        "copied_file_count",
        "copied_files",
    ]
    keys = []
    for row in rows:
        keys.extend(row.keys())
    return [key for key in preferred if key in keys]


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_workbook(path: Path, best_rows: list[dict[str, Any]], all_rows: list[dict[str, Any]]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    workbook = Workbook()
    sheets = [
        ("best_per_model", best_rows),
        ("all_filtered_runs", all_rows),
    ]
    first = True
    for title, rows in sheets:
        worksheet = workbook.active if first else workbook.create_sheet()
        first = False
        worksheet.title = title
        fields = fieldnames_for(rows)
        worksheet.append(fields)
        for row in rows:
            worksheet.append([row.get(field, "") for field in fields])
        worksheet.freeze_panes = "A2"
        for index, field in enumerate(fields, start=1):
            width = min(max(len(field) + 2, 12), 70)
            for cell in worksheet[get_column_letter(index)][1:30]:
                width = min(max(width, len(str(cell.value or "")) + 2), 70)
            worksheet.column_dimensions[get_column_letter(index)].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return True


def build_manifest(
    args: argparse.Namespace,
    metric: str,
    direction: str,
    all_rows: list[dict[str, Any]],
    best_rows: list[dict[str, Any]],
    artifacts: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "filters": {
            "task": normalize_task(args.task),
            "dataset": args.dataset,
            "models": args.model or [],
            "seq_len": args.seq_len,
            "label_len": args.label_len,
            "pred_len": args.pred_len,
            "setting_contains": args.setting_contains,
        },
        "ranking": {"metric": metric, "direction": direction},
        "counts": {"filtered_runs": len(all_rows), "best_models": len(best_rows)},
        "artifacts": artifacts,
    }


def group_key(row: dict[str, Any]) -> tuple[str, str]:
    task = safe_name(str(row.get("task") or "unknown_task"))
    dataset = safe_name(str(row.get("dataset") or "unknown_dataset"))
    return task, dataset


def group_rows(rows: list[dict[str, Any]]) -> dict[tuple[str, str], list[dict[str, Any]]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(group_key(row), []).append(row)
    return grouped


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--results", default="results", help="Path to InduTS-SS results directory.")
    parser.add_argument("--output-dir", default="report", help="Root report directory for grouped Excel files and copied artifacts.")
    parser.add_argument("--task", default="", help="Task filter: SS_task, soft_sensor, LSF_task, or short_term_forecasting.")
    parser.add_argument("--dataset", default="", help="Dataset filter such as DC, SRU, Ironmaking, MP, or PPGAS.")
    parser.add_argument("--model", action="append", help="Model name to include. Repeat for multiple models.")
    parser.add_argument("--seq-len", type=int, default=None, help="Filter by seq_len parsed from setting name.")
    parser.add_argument("--label-len", type=int, default=None, help="Filter by label_len parsed from setting name.")
    parser.add_argument("--pred-len", type=int, default=None, help="Filter by pred_len parsed from setting name.")
    parser.add_argument("--setting-contains", default="", help="Require this substring in the setting directory name.")
    parser.add_argument("--metric", default="mse", choices=METRIC_NAMES + ["corr", "r2"], help="Metric used to choose the best run per model.")
    parser.add_argument("--direction", default="auto", choices=["auto", "min", "max"], help="Ranking direction.")
    parser.add_argument("--no-checkpoints", action="store_true", help="Do not copy model parameter checkpoint files.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    repo_root = Path.cwd()
    metric = "corr_or_r2" if args.metric in {"corr", "r2"} else args.metric
    direction = metric_direction(metric, args.direction)

    all_rows = [row for row in discover_runs(Path(args.results)) if matches_filter(row, args)]
    grouped = group_rows(all_rows)
    output_root = Path(args.output_dir)
    output_root.mkdir(parents=True, exist_ok=True)
    report_index: list[dict[str, Any]] = []
    missing: list[dict[str, Any]] = []
    total_best_rows = 0

    for (task, dataset), group_all_rows in sorted(grouped.items()):
        group_best_rows = select_best(group_all_rows, metric, direction)
        group_output_dir = output_root / task / dataset
        group_output_dir.mkdir(parents=True, exist_ok=True)
        artifacts = collect_artifacts(repo_root, group_best_rows, group_output_dir, not args.no_checkpoints)
        total_best_rows += len(group_best_rows)

        workbook_path = group_output_dir / "best_results.xlsx"
        wrote_xlsx = write_workbook(workbook_path, group_best_rows, group_all_rows)
        if not wrote_xlsx:
            write_csv(group_output_dir / "best_per_model.csv", group_best_rows, fieldnames_for(group_best_rows))
            write_csv(group_output_dir / "all_filtered_runs.csv", group_all_rows, fieldnames_for(group_all_rows))

        manifest = build_manifest(args, metric, direction, group_all_rows, group_best_rows, artifacts)
        manifest["group"] = {"task": task, "dataset": dataset}
        manifest["outputs"] = {
            "workbook": str(workbook_path) if wrote_xlsx else "",
            "best_csv": "" if wrote_xlsx else str(group_output_dir / "best_per_model.csv"),
            "all_runs_csv": "" if wrote_xlsx else str(group_output_dir / "all_filtered_runs.csv"),
            "manifest": str(group_output_dir / "manifest.json"),
        }
        with (group_output_dir / "manifest.json").open("w", encoding="utf-8") as handle:
            json.dump(manifest, handle, indent=2, ensure_ascii=False)

        report_index.append(
            {
                "task": task,
                "dataset": dataset,
                "filtered_runs": len(group_all_rows),
                "best_models": len(group_best_rows),
                "workbook": str(workbook_path) if wrote_xlsx else "",
                "manifest": str(group_output_dir / "manifest.json"),
                "artifact_root": str(group_output_dir / "best_artifacts"),
            }
        )
        missing.extend(
            item for item in artifacts
            if item["missing_checkpoint"] or item["missing_prediction"] or item["missing_truth"]
        )

    with (output_root / "index.json").open("w", encoding="utf-8") as handle:
        json.dump(
            {
                "created_at": datetime.now().isoformat(timespec="seconds"),
                "ranking": {"metric": metric, "direction": direction},
                "filtered_runs": len(all_rows),
                "groups": report_index,
            },
            handle,
            indent=2,
            ensure_ascii=False,
        )

    print(f"Filtered runs: {len(all_rows)}")
    print(f"Groups: {len(report_index)}")
    print(f"Best model rows: {total_best_rows}")
    print(f"Ranking: {metric} ({direction})")
    print(f"Report root: {output_root}")
    print(f"Index: {output_root / 'index.json'}")
    for item in report_index:
        workbook = item["workbook"] or str(Path(item["manifest"]).with_name("best_per_model.csv"))
        print(f"- {item['task']}/{item['dataset']}: {workbook}")
    if missing:
        print("Warnings:")
        for item in missing:
            flags = [
                name for name in ("missing_checkpoint", "missing_prediction", "missing_truth")
                if item[name]
            ]
            print(f"- {item['model']}: {', '.join(flags)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
